import streamlit as st
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings, HuggingFaceInstructEmbeddings
from langchain.vectorstores import FAISS
from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from htmlTemplates import css, bot_template, user_template
from langchain.llms import HuggingFaceHub
import docx
import pandas as pd
import openpyxl
import pytesseract
from PIL import Image



def extract_text_from_docx(docx_file):
    doc = docx.Document(docx_file)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text

def extract_text_from_csv(csv_file):
    df = pd.read_csv(csv_file)
    text = df.to_string(index=False)
    return text

def extract_text_from_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    text = ""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    text += str(cell.value) + "\n"
    return text

def get_pdf_text(pdf_docs):
    text = ""
    for pdf in pdf_docs:
        pdf_reader = PdfReader(pdf)
        for page in pdf_reader.pages:
            text += page.extract_text()
    return text

def get_text_chunks(text, chunk_size=1000):
    chunks = []
    start = 0
    end = chunk_size
    while start < len(text):
        if end >= len(text):
            end = len(text)
        chunk = text[start:end]
        chunks.append(chunk)
        start = end
        end += chunk_size
    return chunks

def get_vectorstore(text_chunks):
    embeddings = OpenAIEmbeddings()
    # embeddings = HuggingFaceInstructEmbeddings(model_name="hkunlp/instructor-xl")
    vectorstore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
    return vectorstore


def get_conversation_chain(vectorstore):
    llm = ChatOpenAI()
    # llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})

    memory = ConversationBufferMemory(
        memory_key='chat_history', return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain


def handle_userinput(user_question):
    response = st.session_state.conversation({'question': user_question})
    st.session_state.chat_history = response['chat_history']

    for i, message in enumerate(st.session_state.chat_history):
        if i % 2 == 0:
            st.write(user_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
        else:
            st.write(bot_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)

def main():
    load_dotenv()
    st.set_page_config(page_title="Chat with multiple documents",
                       page_icon=":books:")
    st.write(css, unsafe_allow_html=True)

    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None

    st.header("Chat with multiple documents :books:")
    user_question = st.text_input("Ask a question about your documents:")
    if user_question:
        handle_userinput(user_question)

    with st.sidebar:
        st.subheader("Your documents")
        pdf_docs = st.file_uploader(
            "Upload your PDFs here", accept_multiple_files=True, type=['pdf'])
        word_docs = st.file_uploader(
            "Upload Word files here", accept_multiple_files=True, type=['docx'])
        csv_docs = st.file_uploader(
            "Upload CSV files here", accept_multiple_files=True, type=['csv'])
        excel_docs = st.file_uploader(
            "Upload Excel files here", accept_multiple_files=True, type=['xlsx'])
        image_docs = st.file_uploader(
        "Upload images here", accept_multiple_files=True, type=['jpg', 'png', 'jpeg', 'gif'])

        if st.button("Process"):
            with st.spinner("Processing"):
                extracted_text = ""
                # Combine all uploaded files of different types into a single list
                all_docs = []
                if pdf_docs:
                    all_docs.extend(pdf_docs)
                if word_docs:
                    all_docs.extend(word_docs)
                if csv_docs:
                    all_docs.extend(csv_docs)
                if excel_docs:
                    all_docs.extend(excel_docs)
                if image_docs:
            # Extract text from uploaded images using OCR
                    for image in image_docs:
                        image = Image.open(image)
                        image_text = pytesseract.image_to_string(image)
                        extracted_text += image_text
                # Process the combined list of documents
                if all_docs:
                    for doc in all_docs:
                        if doc.name.endswith('.pdf'):
                            # Extract text from PDF files using your existing function
                            pdf_text = get_pdf_text([doc])
                            extracted_text += pdf_text
                        elif doc.name.endswith('.docx'):
                            # Extract text from Word files
                            docx_text = extract_text_from_docx(doc)
                            extracted_text += docx_text
                        elif doc.name.endswith('.csv'):
                            # Extract text from CSV files
                            csv_text = extract_text_from_csv(doc)
                            extracted_text += csv_text
                        elif doc.name.endswith('.xlsx'):
                            # Extract text from Excel files
                            excel_text = extract_text_from_excel(doc)
                            extracted_text += excel_text

                # get the text chunks
                text_chunks = get_text_chunks(extracted_text, chunk_size=1000)
                # create vector store
                vectorstore = get_vectorstore(text_chunks)

                # create conversation chain
                st.session_state.conversation = get_conversation_chain(
                    vectorstore)


if __name__ == '__main__':
    main()



